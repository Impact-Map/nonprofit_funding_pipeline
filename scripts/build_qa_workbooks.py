"""Generate non-technical QA workbooks from the lightweight pipeline outputs.

The lightweight analytic outputs (parquet, CSV with raw column names, JSON
manifests) are appropriate for code-driven analysis but not for handing to a
non-technical client for spot-check QA. This script reads those outputs and
produces two formatted Excel workbooks designed for human review:

  Headline_Summary.xlsx    main QA artifact - panel/FY headlines, top
                           agencies, top listings, top recipients per panel,
                           shift-share, caveats, glossary
  Recipient_Lookup.xlsx    per-recipient lookup table for "is org X in the
                           in-scope set" questions, plus the top excluded
                           recipients with plain-English exclusion reasons

Both workbooks use friendly column names, dollar formatting, percent
formatting, banded rows, and frozen headers. The output is the file you
email to a non-technical reviewer alongside the methodology PDF.

Re-run after any methodology change:

    python3 scripts/build_qa_workbooks.py

Outputs land in exhibits/lightweight/qa_for_client/.
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import urllib.parse
from pathlib import Path
from typing import Iterable

import pandas as pd
import yaml

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src import config  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
LOG = logging.getLogger("qa_workbooks")


# ---------------------------------------------------------------------------
# Reference data
# ---------------------------------------------------------------------------

# Two-letter US state code -> human-readable state name. Reviewer-friendly.
STATE_NAMES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "DC": "District of Columbia", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii",
    "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine",
    "MD": "Maryland", "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota",
    "MS": "Mississippi", "MO": "Missouri", "MT": "Montana", "NE": "Nebraska",
    "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico",
    "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island",
    "SC": "South Carolina", "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas",
    "UT": "Utah", "VT": "Vermont", "VA": "Virginia", "WA": "Washington",
    "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming",
    "PR": "Puerto Rico", "VI": "U.S. Virgin Islands", "GU": "Guam",
    "AS": "American Samoa", "MP": "Northern Mariana Islands",
}

PANELS = [
    ("core", "Topline Core (501c3)"),
    ("educational", "Educational"),
    ("hospital", "Hospital"),
    ("international", "International"),
]

PANEL_FILES = {
    "Total": "Total",
    "Topline Core (501c3)": "Topline_Core",
    "Educational": "Educational",
    "Hospital": "Hospital",
    "International": "International",
}


def state_name(code: str | None) -> str:
    if not code or not isinstance(code, str):
        return ""
    return STATE_NAMES.get(code.strip().upper(), code)


def fmt_business_types(s: str | None) -> str:
    """Render a business_types_set string ('MX', 'MEK') as readable list."""
    if not s or not isinstance(s, str):
        return ""
    return ", ".join(sorted(s))


def usaspending_search_url(name: str | None, uei: str | None) -> str:
    """USAspending keyword-search URL. Use recipient name when available
    (more readable; the recipient profile is the first result), fall back
    to UEI which is unique and always finds exactly that recipient."""
    keyword = (name or "").strip() or (uei or "").strip()
    if not keyword:
        return ""
    return f"https://www.usaspending.gov/keyword_search/{urllib.parse.quote(keyword)}"


# ---------------------------------------------------------------------------
# Plain-English exclusion reasons
# ---------------------------------------------------------------------------

def load_business_types_yaml() -> dict:
    path = config.REFERENCE_LISTS / "business_types_lightweight.yaml"
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def translate_exclusion_reason(reason: str, code_to_label: dict[str, str]) -> str:
    """Turn 'disqualifying_cotag:Q+X' into 'Tagged as For-Profit + Other'."""
    if not reason or not isinstance(reason, str):
        return ""
    if reason == "no_M_tag":
        return "No 501(c)(3) tag (M) was found on any of this recipient's transactions."
    if reason.startswith("disqualifying_cotag:"):
        codes = reason.split(":", 1)[1].split("+")
        labels = [code_to_label.get(c, f"code {c}") for c in codes]
        if len(labels) == 1:
            return f"Tagged as {labels[0]} (mutually exclusive with 501(c)(3) status)."
        return ("Tagged as " + ", ".join(labels[:-1]) + " and " + labels[-1] +
                " (mutually exclusive with 501(c)(3) status).")
    return reason


# ---------------------------------------------------------------------------
# CFDA listing title lookup
# ---------------------------------------------------------------------------

def load_listing_titles() -> dict[str, str]:
    """Build a (assistance_listing_number -> assistance_listing_title) map by
    sampling the most recent FY's interim parquet. The lightweight analytic
    table doesn't carry the title, but the raw parquet does."""
    by_listing: dict[str, str] = {}
    for fy in (2025, 2024, 2023, 2022):
        p = config.INTERIM / f"transactions_fy{fy}.parquet"
        if not p.exists():
            continue
        try:
            df = pd.read_parquet(p, columns=["assistance_listing_number", "assistance_listing_title"])
            df = df.dropna(subset=["assistance_listing_number"])
            for k, v in df.drop_duplicates("assistance_listing_number").set_index(
                "assistance_listing_number"
            )["assistance_listing_title"].items():
                if k not in by_listing and v and isinstance(v, str):
                    by_listing[k] = v
        except Exception:
            continue
        if len(by_listing) > 500:
            break
    return by_listing


# ---------------------------------------------------------------------------
# Excel formatting helpers
# ---------------------------------------------------------------------------

DOLLAR_FMT = '"$"#,##0'
DOLLAR_FMT_DECIMAL = '"$"#,##0.00'
PCT_FMT = "0.0%"


def write_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame,
                dollar_columns: Iterable[str] = (),
                percent_columns: Iterable[str] = (),
                col_widths: dict[str, int] | None = None,
                freeze_header: bool = True,
                add_filter: bool = True,
                title: str | None = None) -> None:
    """Write a DataFrame to a sheet with friendly formatting."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    safe_name = sheet_name[:31]  # Excel sheet name limit

    start_row = 1
    if title:
        # Reserve row 1 for the title; data starts row 3 (1-based, openpyxl).
        df.to_excel(writer, sheet_name=safe_name, index=False, startrow=2)
        ws = writer.sheets[safe_name]
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F4E78")
        start_row = 3
    else:
        df.to_excel(writer, sheet_name=safe_name, index=False)
        ws = writer.sheets[safe_name]

    # Header row formatting
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Number formatting on data rows
    dollar_idx = [list(df.columns).index(c) + 1 for c in dollar_columns if c in df.columns]
    pct_idx = [list(df.columns).index(c) + 1 for c in percent_columns if c in df.columns]
    for r in range(start_row + 1, start_row + 1 + len(df)):
        for c in dollar_idx:
            ws.cell(row=r, column=c).number_format = DOLLAR_FMT
        for c in pct_idx:
            ws.cell(row=r, column=c).number_format = PCT_FMT

    # Column widths
    default_widths = {col: 18 for col in df.columns}
    default_widths.update(col_widths or {})
    for col_idx, column in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = default_widths.get(column, 18)

    # Freeze header
    if freeze_header:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto-filter
    if add_filter and len(df) > 0:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A{start_row}:{last_col}{start_row + len(df)}"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def sheet_summary(txn: pd.DataFrame) -> pd.DataFrame:
    """Headline: per-FY total $, real $, recipients, transactions."""
    g = txn.groupby("fy", dropna=False).agg(
        total_obligations_nominal=("federal_action_obligation", "sum"),
        total_obligations_real_FY25=("federal_action_obligation_real", "sum"),
        unique_recipients=("recipient_uei", lambda s: s.nunique()),
        transactions=("transaction_id", "count"),
    ).reset_index()
    g.columns = ["Fiscal Year", "Total Obligations (nominal)",
                 "Total Obligations (FY25 real $)", "Unique Recipients",
                 "Transactions"]
    return g


def sheet_panel_by_fy(txn: pd.DataFrame) -> pd.DataFrame:
    """One row per panel × FY."""
    rows = []
    for cat, label in PANELS:
        sub = txn[txn["recipient_category"] == cat]
        for fy in sorted(txn["fy"].dropna().unique()):
            s = sub[sub["fy"] == fy]
            rows.append({
                "Panel": label,
                "Fiscal Year": int(fy),
                "Obligations (nominal)": s["federal_action_obligation"].sum(),
                "Obligations (FY25 real $)": s["federal_action_obligation_real"].sum(),
                "Unique Recipients": s["recipient_uei"].nunique(),
                "Transactions": len(s),
            })
    return pd.DataFrame(rows)


def sheet_top_agencies(txn: pd.DataFrame, panel: str, top_n: int = 25) -> pd.DataFrame:
    sub = txn if panel == "all" else txn[txn["recipient_category"] == panel]
    g = (sub.groupby(["fy", "awarding_agency"], dropna=False, observed=False)
            ["federal_action_obligation"].sum().reset_index())
    g["rank"] = g.groupby("fy")["federal_action_obligation"].rank(method="dense", ascending=False)
    g = g[g["rank"] <= top_n].sort_values(["fy", "rank"]).reset_index(drop=True)
    g.columns = ["Fiscal Year", "Awarding Agency", "Obligations", "Rank within FY"]
    g["Rank within FY"] = g["Rank within FY"].astype(int)
    return g


def sheet_top_listings(txn: pd.DataFrame, panel: str,
                       title_lookup: dict[str, str], top_n: int = 25) -> pd.DataFrame:
    sub = txn if panel == "all" else txn[txn["recipient_category"] == panel]
    g = (sub.groupby(["fy", "assistance_listing_number"], dropna=False, observed=False)
            ["federal_action_obligation"].sum().reset_index())
    g["rank"] = g.groupby("fy")["federal_action_obligation"].rank(method="dense", ascending=False)
    g = g[g["rank"] <= top_n].sort_values(["fy", "rank"]).reset_index(drop=True)
    g["Program Title"] = g["assistance_listing_number"].map(lambda v: title_lookup.get(v, ""))
    g = g[["fy", "assistance_listing_number", "Program Title", "federal_action_obligation", "rank"]]
    g.columns = ["Fiscal Year", "CFDA #", "Program Title", "Obligations", "Rank within FY"]
    g["Rank within FY"] = g["Rank within FY"].astype(int)
    return g


def sheet_top_recipients(txn: pd.DataFrame, panel: str,
                         names_by_uei: pd.Series,
                         outlay_by_uei: pd.Series | None = None,
                         top_n: int = 50) -> pd.DataFrame:
    """The lightweight analytic txn table doesn't carry recipient_name; pass
    in a name lookup keyed by UEI so we can resolve."""
    sub = txn if panel == "all" else txn[txn["recipient_category"] == panel]
    if sub.empty:
        return pd.DataFrame(columns=[
            "Recipient Name", "State", "Panel Sub-cut", "Total Obligations (FY22-FY25)",
            "FY22", "FY23", "FY24", "FY25",
            "USAspending Search URL", "Reviewer Notes",
        ])
    by_uei = (sub.groupby("recipient_uei")
                .agg(total=("federal_action_obligation", "sum"))
                .reset_index()
                .sort_values("total", ascending=False)
                .head(top_n))
    fy_pivot = (sub.groupby(["recipient_uei", "fy"])["federal_action_obligation"]
                  .sum().unstack(fill_value=0))
    for fy in (2022, 2023, 2024, 2025):
        if fy not in fy_pivot.columns:
            fy_pivot[fy] = 0
    last = (sub.sort_values("action_date", ascending=False)
              .drop_duplicates("recipient_uei").set_index("recipient_uei"))
    out = by_uei.merge(fy_pivot, left_on="recipient_uei", right_index=True, how="left")
    out["Recipient Name"] = out["recipient_uei"].map(names_by_uei).fillna("")
    out["State"] = out["recipient_uei"].map(last["recipient_state"]).map(state_name)
    out["Panel Sub-cut"] = out["recipient_uei"].map(last["recipient_subcategory"]).fillna("")
    if outlay_by_uei is not None:
        out["Total Cumulative Outlays"] = out["recipient_uei"].map(outlay_by_uei).fillna(0)
    else:
        out["Total Cumulative Outlays"] = 0
    out["USAspending Search URL"] = out.apply(
        lambda r: usaspending_search_url(r["Recipient Name"], r["recipient_uei"]),
        axis=1,
    )
    out["Reviewer Notes"] = ""
    out = out.rename(columns={
        "total": "Total Obligations (FY22-FY25)",
        2022: "FY22", 2023: "FY23", 2024: "FY24", 2025: "FY25",
    })
    return out[[
        "Recipient Name", "State", "Panel Sub-cut",
        "Total Obligations (FY22-FY25)",
        "FY22", "FY23", "FY24", "FY25",
        "Total Cumulative Outlays",
        "USAspending Search URL", "Reviewer Notes",
    ]]


def sheet_outlays_by_vintage_fy(awards: pd.DataFrame) -> pd.DataFrame:
    """Cumulative outlays grouped by award-vintage FY × panel.

    Vintage FY = FY of the award's earliest action_date. Outlays in
    USAspending are reported cumulatively over the award's life and are
    attributed to vintage FY here. This produces a 'front-loaded' profile:
    older vintages carry more outlay weight because they've had more time
    to be drawn down.
    """
    if awards.empty or "cumulative_outlay" not in awards.columns:
        return pd.DataFrame()
    a = awards.copy()
    a["cumulative_outlay"] = pd.to_numeric(a["cumulative_outlay"], errors="coerce").fillna(0)
    panels_data = {}
    for cat, label in PANELS:
        sub = a[a["recipient_category"] == cat]
        panels_data[label] = (sub.groupby("vintage_fy")["cumulative_outlay"]
                              .sum().sort_index())
    panels_data["Total"] = a.groupby("vintage_fy")["cumulative_outlay"].sum().sort_index()
    out = pd.DataFrame(panels_data).reset_index()
    out = out.rename(columns={"vintage_fy": "Vintage FY"})
    out["Vintage FY"] = out["Vintage FY"].astype(int)
    n_awards = a.groupby("vintage_fy").size().sort_index()
    out["Awards (count)"] = out["Vintage FY"].map(n_awards).fillna(0).astype(int)
    return out


def sheet_top_recipient_states(txn: pd.DataFrame) -> pd.DataFrame:
    """State-level breakdown of recipient locations, all panels combined."""
    if txn.empty or "recipient_state" not in txn.columns:
        return pd.DataFrame()
    g = (txn.groupby("recipient_state", dropna=False, observed=False)
            .agg(total=("federal_action_obligation", "sum"),
                 unique_recipients=("recipient_uei", lambda s: s.nunique()),
                 transactions=("transaction_id", "count"))
            .reset_index())
    g = g.dropna(subset=["recipient_state"])
    total_dollars = g["total"].sum() or 1
    g["share"] = g["total"] / total_dollars
    g["State Name"] = g["recipient_state"].map(state_name)
    g = g.sort_values("total", ascending=False).reset_index(drop=True)
    g["Rank"] = g.index + 1
    g = g.rename(columns={
        "recipient_state": "State Code",
        "total": "Total Obligations",
        "unique_recipients": "Unique Recipients",
        "transactions": "Transactions",
        "share": "Share of Total",
    })
    return g[["Rank", "State Code", "State Name", "Total Obligations",
              "Share of Total", "Unique Recipients", "Transactions"]]


def sheet_top_pop_states(txn: pd.DataFrame) -> pd.DataFrame:
    """Place-of-performance state breakdown across domestic transactions.

    POP-state is most informative for state-level program analysis (where
    is the federal money actually being spent?). For International, POP
    is country-level so this sheet is filtered to USA POP only.
    """
    if "place_of_performance_state_name" not in txn.columns:
        return pd.DataFrame()
    pop_country = txn.get("place_of_performance_country", pd.Series([""] * len(txn)))
    domestic = txn[pop_country.fillna("").astype(str).str.upper().isin(["USA", ""])]
    if domestic.empty:
        return pd.DataFrame()
    g = (domestic.groupby("place_of_performance_state_name", dropna=False, observed=False)
                 .agg(total=("federal_action_obligation", "sum"),
                      unique_recipients=("recipient_uei", lambda s: s.nunique()),
                      transactions=("transaction_id", "count"))
                 .reset_index())
    g = g.dropna(subset=["place_of_performance_state_name"])
    g = g[g["place_of_performance_state_name"].astype(str).str.strip() != ""]
    total_dollars = g["total"].sum() or 1
    g["share"] = g["total"] / total_dollars
    g = g.sort_values("total", ascending=False).reset_index(drop=True)
    g["Rank"] = g.index + 1
    g = g.rename(columns={
        "place_of_performance_state_name": "POP State",
        "total": "Total Obligations",
        "unique_recipients": "Unique Recipients",
        "transactions": "Transactions",
        "share": "Share of Total",
    })
    return g[["Rank", "POP State", "Total Obligations",
              "Share of Total", "Unique Recipients", "Transactions"]]


def sheet_top_recipient_counties(txn: pd.DataFrame, top_n: int = 100) -> pd.DataFrame:
    """Top-N counties by recipient location."""
    if "recipient_county_name" not in txn.columns or "recipient_state" not in txn.columns:
        return pd.DataFrame()
    g = (txn.groupby(["recipient_state", "recipient_county_name"],
                     dropna=False, observed=False)
            .agg(total=("federal_action_obligation", "sum"),
                 unique_recipients=("recipient_uei", lambda s: s.nunique()),
                 transactions=("transaction_id", "count"))
            .reset_index())
    g = g.dropna(subset=["recipient_state", "recipient_county_name"])
    g = g[g["recipient_county_name"].astype(str).str.strip() != ""]
    g["State"] = g["recipient_state"].map(state_name)
    g = g.sort_values("total", ascending=False).head(top_n).reset_index(drop=True)
    g["Rank"] = g.index + 1
    g = g.rename(columns={
        "recipient_county_name": "County",
        "total": "Total Obligations",
        "unique_recipients": "Unique Recipients",
        "transactions": "Transactions",
    })
    return g[["Rank", "State", "County", "Total Obligations",
              "Unique Recipients", "Transactions"]]


def sheet_yoy_change(txn: pd.DataFrame, top_n: int = 30) -> pd.DataFrame:
    """Largest agency-level FY22 → FY25 changes."""
    f = txn[txn["fy"].isin([2022, 2025])]
    if f.empty:
        return pd.DataFrame()
    # Group by agency × FY, then pivot so agency is the row index and FY
    # is the column. unstack(level="fy") moves FY out of the row index into
    # columns - the prior code unstacked the wrong level and produced an
    # FY-rows-by-agency-columns frame.
    cell = (f.groupby(["awarding_agency", "fy"], dropna=False, observed=False)
              ["federal_action_obligation"].sum())
    pivot = cell.unstack("fy", fill_value=0).reset_index()
    if 2022 not in pivot.columns or 2025 not in pivot.columns:
        return pd.DataFrame()
    pivot = pivot.rename(columns={
        "awarding_agency": "Awarding Agency",
        2022: "FY22 Obligations",
        2025: "FY25 Obligations",
    })[["Awarding Agency", "FY22 Obligations", "FY25 Obligations"]]
    pivot["Change ($)"] = pivot["FY25 Obligations"] - pivot["FY22 Obligations"]
    pivot["Change (%)"] = pivot.apply(
        lambda r: (r["FY25 Obligations"] - r["FY22 Obligations"]) / r["FY22 Obligations"]
        if r["FY22 Obligations"] else None, axis=1,
    )
    pivot = pivot.dropna(subset=["Awarding Agency"])
    return pivot.sort_values("Change ($)", key=abs, ascending=False).head(top_n).reset_index(drop=True)


def sheet_caveats() -> pd.DataFrame:
    rows = [
        ("Recipient identification depends on a single tag",
         "501(c)(3) status is identified by USAspending's 'M' business type code, which is "
         "agency-reported. Some agencies under-tag, over-tag, or use the 'Other' code "
         "alongside M. The methodology accepts this as the cost of running without an IRS-side cross-check."),
        ("No point-in-time IRS verification",
         "If an organization's 501(c)(3) status was revoked during FY22-FY25, the recipient is "
         "still counted because USAspending continues to report the M tag. We cannot detect revocations."),
        ("Educational panel is intentionally narrow",
         "The recipient identification rule excludes universities (public, private, HBCU, TCCU, "
         "Hispanic-serving, Alaska/Native-Hawaiian-serving) and K-12 public school districts. "
         "The Educational panel reports only 501(c)(3) education nonprofits that are NOT institutions "
         "of higher education - charter networks, education foundations, research nonprofits."),
        ("Hospital and Educational panels are heuristic",
         "Recipients are classified by recipient-name regex patterns plus curated assistance-listing "
         "rules (HRSA for FQHCs, NIH for research). False positives and false negatives are inevitable. "
         "A 200-row precision audit per panel is required before quoting these numbers externally."),
        ("FY25 numbers are a floor",
         "USAspending's award data and outlay data both lag real-time by 1-2 quarters. The FY25 "
         "snapshot used here was generated 2026-04-06, so anything obligated or outlaid after that "
         "date is missing. FY25 figures should be read as a lower bound."),
        ("Outlays are vintage-allocated",
         "Outlays in USAspending data are reported cumulatively over the life of an award. "
         "We allocate them to the fiscal year of the award's first action_date. So an award "
         "started in FY22 with outlays continuing into FY25 has all its outlays credited to FY22. "
         "This produces a 'front-loaded' vintage profile."),
        ("Foreign 501(c)(3)s are excluded",
         "Recipients tagged W (Non-domestic Entity) are excluded even if also tagged M. "
         "U.S.-based 501(c)(3) implementing partners with foreign offices that some agency "
         "tagged W are out of scope."),
        ("Sub-awards are out of scope",
         "Federal funds passed through a state government to a 501(c)(3) sub-recipient appear "
         "under the state, not the 501(c)(3). Sub-award (FFATA) data is not analyzed."),
    ]
    return pd.DataFrame(rows, columns=["Caveat", "Plain-English explanation"])


def sheet_glossary() -> pd.DataFrame:
    rows = [
        ("UEI", "Unique Entity Identifier - 12-character code USAspending uses to identify each recipient organization. Replaced DUNS number in April 2022."),
        ("EIN", "Employer Identification Number - the 9-digit IRS tax ID. Not used in this analysis (we identify 501(c)(3) status from USAspending tags, not IRS records)."),
        ("CFDA / Assistance Listing", "A unique identifier for each federal grant program. The CFDA number maps to a program title (e.g., 93.224 = Health Center Program)."),
        ("Obligation", "A binding commitment by the federal government to spend money. Reported per-transaction with a dated action_date."),
        ("Outlay", "An actual disbursement of federal money. Reported by USAspending only at the prime award level (cumulative over the award's life), not per-transaction."),
        ("Fiscal Year (FY)", "The U.S. federal fiscal year runs October 1 to September 30. FY 2024 = Oct 1, 2023 to Sep 30, 2024."),
        ("Vintage FY", "For outlays: the FY of the award's first action. We attribute the award's cumulative life-of-award outlay to its vintage FY."),
        ("Real dollars (FY25)", "Inflation-adjusted using BLS CPI-U so historical FYs are comparable to FY25."),
        ("Panel", "One of four mutually exclusive categories: Topline Core, Educational, Hospital, International. Each transaction is assigned to exactly one panel."),
        ("Topline Core", "The headline 501(c)(3) figure - residual after Educational, Hospital, and International carve-outs."),
        ("In scope / out of scope", "A recipient is 'in scope' if their USAspending business_types_code includes 'M' (501(c)(3)) and excludes any disqualifying co-tag (e.g., for-profit, state government, university)."),
        ("M tag", "USAspending business_types_code letter for 'Nonprofit with 501C3 IRS Status, Other than an Institution of Higher Education'."),
        ("Tagged as Other (X)", "Some agencies enter both 'M' and 'X' (Other) on the same recipient. We tolerate the X co-tag because it doesn't contradict 501(c)(3) status."),
    ]
    return pd.DataFrame(rows, columns=["Term", "Plain-English meaning"])


def sheet_recipient_lookup(txn: pd.DataFrame, names_by_uei: pd.Series,
                           outlay_by_uei: pd.Series | None = None) -> pd.DataFrame:
    """Per-recipient roll-up across FY22-FY25 with searchable columns."""
    if txn.empty:
        return pd.DataFrame()
    fy_pivot = (txn.groupby(["recipient_uei", "fy"])["federal_action_obligation"]
                  .sum().unstack(fill_value=0))
    for fy in (2022, 2023, 2024, 2025):
        if fy not in fy_pivot.columns:
            fy_pivot[fy] = 0
    last = (txn.sort_values("action_date", ascending=False)
              .drop_duplicates("recipient_uei").set_index("recipient_uei"))

    out = pd.DataFrame(index=fy_pivot.index)
    out["Recipient Name"] = pd.Series(out.index, index=out.index).map(names_by_uei).fillna("")
    out["State"] = last["recipient_state"].map(state_name)
    # Geographic context (Tier C - mapping fields). Pulled from the analytic
    # table when present; missing columns degrade silently to empty strings.
    for col, label in [
        ("recipient_city", "City"),
        ("recipient_county_name", "County"),
        ("recipient_zip", "ZIP"),
        ("recipient_cd", "Congressional District"),
    ]:
        out[label] = last[col].fillna("") if col in last.columns else ""
    out["Panel"] = last["recipient_category"].map(dict(PANELS))
    out["Panel Sub-cut"] = last["recipient_subcategory"].fillna("")
    out["Business Types Tagged"] = last["business_types_set"].fillna("").map(fmt_business_types)
    out["FY22 Obligations"] = fy_pivot[2022]
    out["FY23 Obligations"] = fy_pivot[2023]
    out["FY24 Obligations"] = fy_pivot[2024]
    out["FY25 Obligations"] = fy_pivot[2025]
    out["Total FY22-FY25"] = (fy_pivot[2022] + fy_pivot[2023] +
                              fy_pivot[2024] + fy_pivot[2025])
    if outlay_by_uei is not None:
        out["Total Cumulative Outlays"] = (
            pd.Series(out.index, index=out.index).map(outlay_by_uei).fillna(0)
        )
    else:
        out["Total Cumulative Outlays"] = 0
    out["UEI"] = out.index
    out["USAspending Search URL"] = out.apply(
        lambda r: usaspending_search_url(r["Recipient Name"], r["UEI"]), axis=1,
    )
    return out.sort_values("Total FY22-FY25", ascending=False).reset_index(drop=True)


def sheet_top_excluded(rfilter: pd.DataFrame, txn_raw: pd.DataFrame,
                       code_to_label: dict[str, str], top_n: int = 200) -> pd.DataFrame:
    """Top excluded recipients by associated obligated dollars."""
    excluded = rfilter[~rfilter["in_scope"]].copy()
    if excluded.empty:
        return pd.DataFrame()
    txn_raw["federal_action_obligation"] = pd.to_numeric(
        txn_raw["federal_action_obligation"], errors="coerce"
    ).fillna(0)
    obl = txn_raw.groupby("recipient_uei", as_index=False)["federal_action_obligation"].sum()
    excluded = excluded.merge(obl, on="recipient_uei", how="left").fillna({"federal_action_obligation": 0})
    excluded["Why Excluded"] = excluded["exclusion_reason"].map(
        lambda r: translate_exclusion_reason(r, code_to_label)
    )
    excluded["State"] = excluded["recipient_state"].map(state_name)
    excluded["Business Types Tagged"] = excluded["bt_set"].fillna("").map(fmt_business_types)
    excluded["USAspending Search URL"] = excluded.apply(
        lambda r: usaspending_search_url(r.get("recipient_name"), r["recipient_uei"]),
        axis=1,
    )
    excluded = excluded.rename(columns={
        "recipient_name": "Recipient Name",
        "federal_action_obligation": "Total FY22-FY25 Obligations",
        "recipient_uei": "UEI",
    })
    excluded = excluded.sort_values("Total FY22-FY25 Obligations", ascending=False).head(top_n)
    return excluded[[
        "Recipient Name", "State", "Total FY22-FY25 Obligations",
        "Why Excluded", "Business Types Tagged", "UEI", "USAspending Search URL",
    ]].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Workbook composition
# ---------------------------------------------------------------------------

def build_headline_workbook(out_path: Path, txn: pd.DataFrame,
                            awards: pd.DataFrame,
                            names_by_uei: pd.Series,
                            outlay_by_uei: pd.Series,
                            title_lookup: dict[str, str], snapshot_info: dict) -> None:
    LOG.info("Building %s", out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Cover sheet
        cover = pd.DataFrame([
            ("Federal Financial Assistance to 501(c)(3) Organizations, FY22-FY25", ""),
            ("Recipient identification:", "USAspending business_types_code 'M' tag with mutually-exclusive co-tag exclusions"),
            ("Snapshot date (Award Data Archive):", snapshot_info.get("snapshot_date", "")),
            ("Run completed:", snapshot_info.get("run_date", "")),
            ("Methodology version:", snapshot_info.get("methodology_version", "v2 (X-tolerant)")),
            ("", ""),
            ("How to read this workbook:", ""),
            ("1. Summary tab", "Topline numbers per FY"),
            ("2. By Panel x FY tab", "Same numbers split by panel"),
            ("3. Outlays by Vintage FY tab", "Cumulative outlays attributed to the award's first action FY (read the caveats — vintage allocation is front-loaded)"),
            ("4. Top 25 Agencies tabs", "One per panel - which agencies fund each panel?"),
            ("5. Top 25 Programs tabs", "One per panel - which CFDA programs?"),
            ("6. Top 50 Recipients tabs", "One per panel. Use this for spot-check QA: do these orgs match what you'd expect? Includes Total Cumulative Outlays per recipient."),
            ("7. By Recipient State tab", "All 51 states/territories ranked by total dollars and recipient count"),
            ("8. By POP State tab", "Place-of-performance state breakdown (where the work is happening)"),
            ("9. Top Recipient Counties tab", "Top 100 counties by recipient location"),
            ("10. YoY Change tab", "Largest agency-level changes from FY22 to FY25"),
            ("11. Caveats tab", "Plain-English limitations of this analysis"),
            ("12. Glossary tab", "Definitions of any technical terms used"),
            ("", ""),
            ("For QA spot-checks, focus on Top 50 Recipients tabs:", ""),
            ("- Are the named organizations actually 501(c)(3)?", ""),
            ("- Are they assigned to the right panel?", ""),
            ("- Is the dollar magnitude in line with what you know?", ""),
            ("- Click the USAspending Search URL for any recipient to verify", ""),
        ], columns=["", ""])
        write_sheet(writer, "Cover", cover, freeze_header=False, add_filter=False,
                    col_widths={"": 60})

        # Headline sheets
        write_sheet(writer, "Summary", sheet_summary(txn),
                    title="Total federal financial assistance to 501(c)(3) recipients, by FY",
                    dollar_columns=["Total Obligations (nominal)", "Total Obligations (FY25 real $)"],
                    col_widths={"Total Obligations (nominal)": 24, "Total Obligations (FY25 real $)": 24})

        write_sheet(writer, "By Panel x FY", sheet_panel_by_fy(txn),
                    title="Obligations by panel and fiscal year",
                    dollar_columns=["Obligations (nominal)", "Obligations (FY25 real $)"],
                    col_widths={"Panel": 22, "Obligations (nominal)": 22, "Obligations (FY25 real $)": 22})

        # Outlays by award-vintage FY × panel. Vintage allocation produces a
        # front-loaded profile (older vintages have more time to be drawn
        # down); the caveat sheet explains. Skipped if awards table missing.
        df_outlays = sheet_outlays_by_vintage_fy(awards) if not awards.empty else pd.DataFrame()
        if not df_outlays.empty:
            write_sheet(writer, "Outlays by Vintage FY", df_outlays,
                        title=("Cumulative outlays by FY of award vintage. NOTE: outlays are "
                               "cumulative-life-of-award, attributed to the award's first "
                               "action FY. FY22 vintage carries more weight because those awards "
                               "have had longer to draw down. See Caveats sheet."),
                        dollar_columns=[lbl for _, lbl in PANELS] + ["Total"],
                        col_widths={"Vintage FY": 12, "Awards (count)": 14,
                                    **{lbl: 22 for _, lbl in PANELS}, "Total": 22})

        # Top 25 agencies — one per panel
        for cat, label in PANELS:
            df = sheet_top_agencies(txn, cat, top_n=25)
            write_sheet(writer, f"Top Agencies - {label}", df,
                        title=f"Top 25 awarding agencies per FY — {label}",
                        dollar_columns=["Obligations"],
                        col_widths={"Awarding Agency": 38, "Obligations": 18})

        # Top 25 listings — one per panel
        for cat, label in PANELS:
            df = sheet_top_listings(txn, cat, title_lookup, top_n=25)
            write_sheet(writer, f"Top Programs - {label}", df,
                        title=f"Top 25 CFDA programs per FY — {label}",
                        dollar_columns=["Obligations"],
                        col_widths={"CFDA #": 10, "Program Title": 50, "Obligations": 18})

        # Top 50 recipients per panel — the headline QA artifact
        for cat, label in PANELS:
            df = sheet_top_recipients(txn, cat, names_by_uei,
                                      outlay_by_uei=outlay_by_uei, top_n=50)
            write_sheet(writer, f"Top 50 - {label}", df,
                        title=f"Top 50 recipients by FY22-FY25 obligations — {label}",
                        dollar_columns=["Total Obligations (FY22-FY25)", "FY22", "FY23",
                                        "FY24", "FY25", "Total Cumulative Outlays"],
                        col_widths={"Recipient Name": 38, "Panel Sub-cut": 18,
                                    "Total Cumulative Outlays": 22,
                                    "USAspending Search URL": 50, "Reviewer Notes": 36})

        # Geographic breakouts (Tier C - mapping)
        df_states = sheet_top_recipient_states(txn)
        if not df_states.empty:
            write_sheet(writer, "By Recipient State", df_states,
                        title="Recipients ranked by state — total dollars and unique organizations",
                        dollar_columns=["Total Obligations"],
                        percent_columns=["Share of Total"],
                        col_widths={"State Code": 12, "State Name": 24,
                                    "Total Obligations": 22, "Share of Total": 14,
                                    "Unique Recipients": 18, "Transactions": 14, "Rank": 8})
        df_pop = sheet_top_pop_states(txn)
        if not df_pop.empty:
            write_sheet(writer, "By POP State (domestic)", df_pop,
                        title="Place-of-performance state breakdown — where the work is happening (USA POP only)",
                        dollar_columns=["Total Obligations"],
                        percent_columns=["Share of Total"],
                        col_widths={"POP State": 30, "Total Obligations": 22,
                                    "Share of Total": 14, "Unique Recipients": 18,
                                    "Transactions": 14, "Rank": 8})
        df_counties = sheet_top_recipient_counties(txn, top_n=100)
        if not df_counties.empty:
            write_sheet(writer, "Top 100 Recipient Counties", df_counties,
                        title="Top 100 counties by total recipient obligations",
                        dollar_columns=["Total Obligations"],
                        col_widths={"State": 22, "County": 30,
                                    "Total Obligations": 22, "Unique Recipients": 18,
                                    "Transactions": 14, "Rank": 8})

        write_sheet(writer, "YoY Change FY22-FY25", sheet_yoy_change(txn),
                    title="Largest agency-level changes from FY22 to FY25",
                    dollar_columns=["FY22 Obligations", "FY25 Obligations", "Change ($)"],
                    percent_columns=["Change (%)"],
                    col_widths={"Awarding Agency": 38})

        write_sheet(writer, "Caveats", sheet_caveats(),
                    title="Limitations a reviewer should be aware of",
                    col_widths={"Caveat": 38, "Plain-English explanation": 90}, add_filter=False)

        write_sheet(writer, "Glossary", sheet_glossary(),
                    title="Plain-English definitions of technical terms",
                    col_widths={"Term": 22, "Plain-English meaning": 90}, add_filter=False)


def build_recipient_lookup_workbook(out_path: Path, txn: pd.DataFrame,
                                    rfilter: pd.DataFrame, txn_raw: pd.DataFrame,
                                    names_by_uei: pd.Series,
                                    outlay_by_uei: pd.Series,
                                    code_to_label: dict[str, str],
                                    snapshot_info: dict) -> None:
    LOG.info("Building %s", out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        cover = pd.DataFrame([
            ("Recipient lookup — Federal Financial Assistance to 501(c)(3), FY22-FY25", ""),
            ("Snapshot date:", snapshot_info.get("snapshot_date", "")),
            ("", ""),
            ("Use the In-Scope Recipients tab to look up specific organizations.", ""),
            ("Filter or sort by Recipient Name, State, or Panel.", ""),
            ("", ""),
            ("If an organization you expected isn't there, check the Top 200 Excluded tab", ""),
            ("for a plain-English exclusion reason.", ""),
        ], columns=["", ""])
        write_sheet(writer, "Cover", cover, freeze_header=False, add_filter=False,
                    col_widths={"": 70})

        in_scope = sheet_recipient_lookup(txn, names_by_uei,
                                          outlay_by_uei=outlay_by_uei)
        write_sheet(writer, "In-Scope Recipients", in_scope,
                    title="All 501(c)(3) recipients in scope (FY22-FY25 union)",
                    dollar_columns=["FY22 Obligations", "FY23 Obligations",
                                    "FY24 Obligations", "FY25 Obligations",
                                    "Total FY22-FY25", "Total Cumulative Outlays"],
                    col_widths={"Recipient Name": 40, "State": 16,
                                "City": 22, "County": 22, "ZIP": 12,
                                "Congressional District": 16,
                                "Panel": 22, "Panel Sub-cut": 18,
                                "Business Types Tagged": 18,
                                "Total Cumulative Outlays": 22, "UEI": 14,
                                "USAspending Search URL": 50})

        excluded = sheet_top_excluded(rfilter, txn_raw, code_to_label, top_n=200)
        write_sheet(writer, "Top 200 Excluded", excluded,
                    title="Largest excluded recipients (plain-English exclusion reason)",
                    dollar_columns=["Total FY22-FY25 Obligations"],
                    col_widths={"Recipient Name": 40, "State": 18,
                                "Why Excluded": 70, "Business Types Tagged": 22,
                                "UEI": 14, "USAspending Search URL": 50})


# ---------------------------------------------------------------------------
# Snapshot info from manifests
# ---------------------------------------------------------------------------

def latest_lightweight_manifest() -> dict:
    """Pick the most recent manifest that ran the lightweight pipeline."""
    out = {}
    for p in sorted(config.MANIFESTS.glob("run-*.json"), reverse=True):
        try:
            data = json.loads(p.read_text())
        except Exception:
            continue
        notes = " ".join(data.get("notes") or [])
        if "lightweight" in (notes or "").lower():
            continue
        if data.get("category_stats") and "rows_total" in (data["category_stats"] or {}):
            out["run_date"] = data.get("started_at", "")[:10]
            for d in data.get("usaspending_downloads", []) or []:
                if "snapshot_dates" in d and d["snapshot_dates"]:
                    out["snapshot_date"] = d["snapshot_dates"][0]
                    break
            break
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    p.add_argument("--out-dir", default=str(config.EXHIBITS / "lightweight" / "qa_for_client"))
    args = p.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    LOG.info("Loading lightweight analytic outputs...")
    txn = pd.read_parquet(config.PROCESSED / "assistance_txn_501c3_lightweight.parquet")
    rfilter = pd.read_parquet(config.PROCESSED / "recipient_filter_lightweight.parquet")
    awards_path = config.PROCESSED / "assistance_awards_501c3_lightweight.parquet"
    if awards_path.exists():
        awards = pd.read_parquet(awards_path)
        # Outlay lookup keyed by UEI: per-recipient sum across all their awards.
        awards["cumulative_outlay"] = pd.to_numeric(
            awards["cumulative_outlay"], errors="coerce"
        ).fillna(0)
        outlay_by_uei = (awards.groupby("recipient_uei")["cumulative_outlay"]
                                .sum())
        LOG.info("Loaded %d in-scope transactions, %d recipients in filter table, %d awards",
                 len(txn), len(rfilter), len(awards))
    else:
        awards = pd.DataFrame()
        outlay_by_uei = pd.Series(dtype=float)
        LOG.warning("Awards parquet missing at %s; outlay sheets will be skipped", awards_path)

    # The lightweight analytic table doesn't carry recipient_name (the
    # M-rule operates on UEI; names live in recipient_filter and the raw
    # transactions). Build a UEI -> name lookup.
    names_by_uei = (rfilter.dropna(subset=["recipient_name"])
                          .drop_duplicates("recipient_uei")
                          .set_index("recipient_uei")["recipient_name"])
    LOG.info("Built UEI -> name lookup: %d entries", len(names_by_uei))

    LOG.info("Loading interim transactions for CFDA-title lookup and obligation aggregation...")
    title_lookup = load_listing_titles()
    LOG.info("CFDA title lookup: %d listings", len(title_lookup))

    # For the excluded-recipient sheet we need raw obligations across all
    # recipients (not just in-scope). Sample one column-projected read.
    txn_raw = pd.concat([
        pd.read_parquet(config.INTERIM / f"transactions_fy{fy}.parquet",
                        columns=["recipient_uei", "federal_action_obligation"])
        for fy in (2022, 2023, 2024, 2025)
        if (config.INTERIM / f"transactions_fy{fy}.parquet").exists()
    ], ignore_index=True)

    # Plain-English code labels for exclusion reasons
    bt_yaml = load_business_types_yaml()
    code_to_label = dict(bt_yaml.get("excluded_codes") or {})

    snapshot_info = latest_lightweight_manifest()

    build_headline_workbook(out_dir / "Headline_Summary.xlsx", txn, awards,
                            names_by_uei, outlay_by_uei,
                            title_lookup, snapshot_info)
    build_recipient_lookup_workbook(out_dir / "Recipient_Lookup.xlsx",
                                    txn, rfilter, txn_raw, names_by_uei,
                                    outlay_by_uei, code_to_label, snapshot_info)

    LOG.info("Done. Files in %s", out_dir)
    for p in sorted(out_dir.glob("*.xlsx")):
        size_mb = p.stat().st_size / 1e6
        LOG.info("  %s (%.1f MB)", p.name, size_mb)


if __name__ == "__main__":
    main()
